# ---------- FUNCIONES GENERALES ----------
def ejecutar_sql(sql, params=(), commit=False):
    con = conectar()
    cur = con.cursor()
    cur.execute(sql, params)
    if commit:
        con.commit()
    res = cur.fetchall() if cur.description else None
    con.close()
    return res

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from tkinter import filedialog

def ajustar_columnas(ws, headers):
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

def reporte_general_excel():
    ruta = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")],
        initialfile="reporte_general.xlsx"
    )
    if not ruta:
        return

    wb = Workbook()

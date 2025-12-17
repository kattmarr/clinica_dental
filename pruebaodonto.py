import customtkinter as ctk
from tkinter import ttk, messagebox
import mysql.connector
from datetime import date
from tkcalendar import Calendar
from PIL import Image


# ---------- CONEXION A MYSQL ----------
def conectar():
    return mysql.connector.connect(
        host="localhost",
        user="root",
        password="",
        database="clinica_dental"
    )

# ---------- FUNCIONES GENERALES ----------
def ejecutar_sql(sql, params=(), commit=False):
    con = conectar()
    cur = con.cursor()
    cur.execute(sql, params)
    if commit:
        con.commit()
    res = cur.fetchall() if cur.description else None
    con.close()
    return res

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from tkinter import filedialog

def ajustar_columnas(ws, headers):
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

def reporte_general_excel():
    ruta = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")],
        initialfile="reporte_general.xlsx"
    )
    if not ruta:
        return

    wb = Workbook()

    # ---------- HOJA: PACIENTES ----------
    ws = wb.active
    ws.title = "Pacientes"

    headers_p = ["ID", "Nombre", "Apellido", "Telefono", "Direccion", "Email"]
    ws.append(headers_p)

    filas_p = ejecutar_sql("""
        SELECT pac_id, nombre, apellido, telefono, direccion, email
        FROM paciente
        ORDER BY pac_id DESC
    """) or []

    for fila in filas_p:
        ws.append(list(fila))

    ajustar_columnas(ws, headers_p)

    # ---------- HOJA: TURNOS ----------
    ws = wb.create_sheet("Turnos")

    headers_t = ["ID", "Paciente", "Odontologo", "Fecha", "Hora", "Motivo", "Estado"]
    ws.append(headers_t)

    filas_t = ejecutar_sql("""
        SELECT t.tur_id,
               CONCAT(p.nombre,' ',p.apellido) AS paciente,
               CONCAT(o.nombre,' ',o.apellido) AS odontologo,
               t.fecha, t.hora, t.motivo, t.estado
        FROM turno t
        JOIN paciente p ON p.pac_id = t.pac_id
        JOIN odontologo o ON o.odo_id = t.odo_id
        ORDER BY t.fecha DESC, t.hora DESC
    """) or []

    for fila in filas_t:
        ws.append(list(fila))

    ajustar_columnas(ws, headers_t)

    # ---------- HOJA: PAGOS ----------
    ws = wb.create_sheet("Pagos")

    headers_pg = ["ID", "Paciente", "Fecha", "Monto", "Metodo"]
    ws.append(headers_pg)

    filas_pg = ejecutar_sql("""
        SELECT p.pag_id,
               CONCAT(pa.nombre,' ',pa.apellido) AS paciente,
               p.fecha, p.monto, p.metodo
        FROM pago p
        JOIN paciente pa ON pa.pac_id = p.pac_id
        ORDER BY p.fecha DESC
    """) or []

    for fila in filas_pg:
        ws.append(list(fila))

    ajustar_columnas(ws, headers_pg)

    wb.save(ruta)
    messagebox.showinfo("Reporte", "Reporte general generado correctamente")


# ---------- MODULO PACIENTES ----------
def ventana_pacientes():
    v = ctk.CTkToplevel()
    v.title("Pacientes")
    v.geometry("700x400")

    ctk.CTkLabel(v, text="Gestión de Pacientes", font=("Arial", 20, "bold")).pack(pady=10)
    frame = ctk.CTkFrame(v)
    frame.pack(pady=10)

    campos = {}
    for i, campo in enumerate(["Nombre", "Apellido", "Teléfono", "Dirección", "Email"]):
        ctk.CTkLabel(frame, text=campo+":").grid(row=i, column=0, padx=10, pady=5, sticky="e")
        entrada = ctk.CTkEntry(frame, width=250)
        entrada.grid(row=i, column=1, padx=10, pady=5)
        campos[campo.lower()] = entrada

    def guardar():
        sql = """INSERT INTO paciente (nombre, apellido, telefono, direccion, email)
                 VALUES (%s,%s,%s,%s,%s)"""
        datos = tuple(campos[c].get() for c in campos)
        ejecutar_sql(sql, datos, commit=True)
        messagebox.showinfo("Éxito", "Paciente guardado correctamente")
        cargar_tabla()

    ctk.CTkButton(frame, text="Guardar", command=guardar).grid(row=6, column=0, columnspan=2, pady=10)

    # Tabla
    tabla = ttk.Treeview(v, columns=("ID","Nombre","Apellido","Teléfono"), show="headings")
    for col in ("ID","Nombre","Apellido","Teléfono"):
        tabla.heading(col, text=col)
        tabla.column(col, width=150)
    tabla.pack(pady=10, fill="x")

    def cargar_tabla():
        for f in tabla.get_children():
            tabla.delete(f)
        filas = ejecutar_sql("SELECT pac_id, nombre, apellido, telefono FROM paciente")
        for fila in filas:
            tabla.insert("", "end", values=fila)

    cargar_tabla()

def reporte_pacientes():
    filas = ejecutar_sql("""
        SELECT pac_id, nombre, apellido, telefono, direccion, email
        FROM paciente
        ORDER BY pac_id
    """)

    encabezados = ["ID", "Nombre", "Apellido", "Telefono", "Direccion", "Email"]

    exportar_excel("reporte_pacientes", encabezados, filas)


# ---------- MODULO ODONTOLOGOS ----------
def ventana_odontologos():
    v = ctk.CTkToplevel()
    v.title("Odontólogos")
    v.geometry("600x400")

    ctk.CTkLabel(v, text="Gestión de Odontólogos",
                 font=("Arial", 20, "bold")).pack(pady=10)

    frame = ctk.CTkFrame(v)
    frame.pack(pady=10)

    campos = {}
    for i, campo in enumerate(["Nombre","Apellido","Especialidad","Teléfono","Email"]):
        ctk.CTkLabel(frame, text=campo+":").grid(row=i, column=0, padx=10, pady=5)
        entrada = ctk.CTkEntry(frame, width=250)
        entrada.grid(row=i, column=1, padx=10, pady=5)
        campos[campo.lower()] = entrada

    def guardar():
        sql = """INSERT INTO odontologo (nombre, apellido, especialidad, telefono, email)
                 VALUES (%s,%s,%s,%s,%s)"""
        datos = tuple(campos[c].get() for c in campos)
        ejecutar_sql(sql, datos, commit=True)
        messagebox.showinfo("Éxito", "Odontólogo guardado correctamente")
        cargar_tabla()

    ctk.CTkButton(frame, text="Guardar", command=guardar)\
        .grid(row=6, column=0, columnspan=2, pady=10)

    # Tabla
    tabla = ttk.Treeview(
        v,
        columns=("ID","Nombre","Apellido","Especialidad"),
        show="headings"
    )

    for col in ("ID","Nombre","Apellido","Especialidad"):
        tabla.heading(col, text=col)
        tabla.column(col, width=150)

    tabla.pack(pady=10, fill="x")

    def cargar_tabla():
        for f in tabla.get_children():
            tabla.delete(f)
        filas = ejecutar_sql(
            "SELECT odo_id, nombre, apellido, especialidad FROM odontologo"
        )
        for fila in filas:
            tabla.insert("", "end", values=fila)

    cargar_tabla()


    

# ---------- MODULO TURNOS ----------
def ventana_turnos():
    v = ctk.CTkToplevel()
    v.title("Turnos")
    v.geometry("700x400")
    ctk.CTkLabel(v, text="Gestión de Turnos", font=("Arial", 20, "bold")).pack(pady=10)
    frame = ctk.CTkFrame(v)
    frame.pack(pady=10)

    pacientes = ejecutar_sql("SELECT pac_id, CONCAT(nombre,' ',apellido) FROM paciente")
    odontologos = ejecutar_sql("SELECT odo_id, CONCAT(nombre,' ',apellido) FROM odontologo")

    lista_pac = {f"{n}": i for i,n in pacientes}
    lista_odo = {f"{n}": i for i,n in odontologos}

    cb_pac = ctk.CTkComboBox(frame, values=list(lista_pac.keys()), width=250)
    cb_odo = ctk.CTkComboBox(frame, values=list(lista_odo.keys()), width=250)
    fecha = ctk.CTkEntry(frame, placeholder_text="YYYY-MM-DD")
    hora = ctk.CTkEntry(frame, placeholder_text="HH:MM")
    motivo = ctk.CTkEntry(frame, placeholder_text="Motivo de la consulta", width=250)

    ctk.CTkLabel(frame, text="Paciente:").grid(row=0, column=0, pady=5)
    cb_pac.grid(row=0, column=1)
    ctk.CTkLabel(frame, text="Odontólogo:").grid(row=1, column=0, pady=5)
    cb_odo.grid(row=1, column=1)
    ctk.CTkLabel(frame, text="Fecha:").grid(row=2, column=0)
    fecha.grid(row=2, column=1)
    ctk.CTkLabel(frame, text="Hora:").grid(row=3, column=0)
    hora.grid(row=3, column=1)
    ctk.CTkLabel(frame, text="Motivo:").grid(row=4, column=0)
    motivo.grid(row=4, column=1)

    def guardar_turno():
        sql = """INSERT INTO turno (pac_id, odo_id, fecha, hora, motivo)
                 VALUES (%s,%s,%s,%s,%s)"""
        datos = (lista_pac[cb_pac.get()], lista_odo[cb_odo.get()], fecha.get(), hora.get(), motivo.get())
        ejecutar_sql(sql, datos, commit=True)
        messagebox.showinfo("Éxito", "Turno registrado correctamente")
        cargar_tabla()

    ctk.CTkButton(frame, text="Guardar", command=guardar_turno).grid(row=5, column=0, columnspan=2, pady=10)

    tabla = ttk.Treeview(v, columns=("ID","Paciente","Odontólogo","Fecha","Hora"), show="headings")
    for col in ("ID","Paciente","Odontólogo","Fecha","Hora"):
        tabla.heading(col, text=col)
    tabla.pack(fill="x", pady=10)

    def cargar_tabla():
        for f in tabla.get_children():
            tabla.delete(f)
        filas = ejecutar_sql("""SELECT t.tur_id, p.nombre, o.nombre, t.fecha, t.hora
                             FROM turno t
                             JOIN paciente p ON p.pac_id=t.pac_id
                             JOIN odontologo o ON o.odo_id=t.odo_id""")
        for fila in filas:
            tabla.insert("", "end", values=fila)

    cargar_tabla()
turnos = ejecutar_sql("""
    SELECT 
        t.tur_id,
        CONCAT(p.nombre,' ',p.apellido) AS paciente,
        CONCAT(o.nombre,' ',o.apellido) AS odontologo,
        t.fecha,
        t.hora,
        t.motivo
    FROM turno t
    LEFT JOIN paciente p ON p.pac_id = t.pac_id
    LEFT JOIN odontologo o ON o.odo_id = t.odo_id
""")




# ---------- MODULO TRATAMIENTOS ----------
def ventana_tratamientos():
    v = ctk.CTkToplevel()
    v.title("Tratamientos")
    v.geometry("600x400")
    ctk.CTkLabel(v, text="Gestión de Tratamientos", font=("Arial", 20, "bold")).pack(pady=10)
    frame = ctk.CTkFrame(v)
    frame.pack(pady=10)

    nombre = ctk.CTkEntry(frame, width=250, placeholder_text="Nombre del tratamiento")
    desc = ctk.CTkEntry(frame, width=250, placeholder_text="Descripción")
    precio = ctk.CTkEntry(frame, width=150, placeholder_text="Precio")

    ctk.CTkLabel(frame, text="Nombre:").grid(row=0, column=0)
    nombre.grid(row=0, column=1)
    ctk.CTkLabel(frame, text="Descripción:").grid(row=1, column=0)
    desc.grid(row=1, column=1)
    ctk.CTkLabel(frame, text="Precio:").grid(row=2, column=0)
    precio.grid(row=2, column=1)

    def guardar():
        ejecutar_sql("INSERT INTO tratamiento (nombre, descripcion, precio) VALUES (%s,%s,%s)",
                     (nombre.get(), desc.get(), precio.get()), commit=True)
        messagebox.showinfo("Éxito", "Tratamiento guardado")
        cargar_tabla()

    ctk.CTkButton(frame, text="Guardar", command=guardar).grid(row=3, column=0, columnspan=2, pady=10)

    tabla = ttk.Treeview(v, columns=("ID","Nombre","Precio"), show="headings")
    for col in ("ID","Nombre","Precio"):
        tabla.heading(col, text=col)
    tabla.pack(fill="x", pady=10)

    def cargar_tabla():
        for f in tabla.get_children():
            tabla.delete(f)
        filas = ejecutar_sql("SELECT tra_id, nombre, precio FROM tratamiento")
        for fila in filas:
            tabla.insert("", "end", values=fila)

    cargar_tabla()

# ---------- MODULO HISTORIAL DEL PACIENTE ----------
def ventana_historial_paciente():
    v = ctk.CTkToplevel()
    v.title("Historial del paciente")
    v.geometry("900x500")

    ctk.CTkLabel(v, text="Historial del paciente",
                 font=("Arial", 20, "bold")).pack(pady=10)

    frame_top = ctk.CTkFrame(v)
    frame_top.pack(pady=5)

    pacientes = ejecutar_sql("SELECT pac_id, CONCAT(nombre,' ',apellido) FROM paciente")
    lista_pac = {n: i for i, n in pacientes}

    ctk.CTkLabel(frame_top, text="Paciente:").grid(row=0, column=0, padx=5, pady=5)
    cb_pac = ctk.CTkComboBox(frame_top, values=list(lista_pac.keys()), width=300)
    cb_pac.grid(row=0, column=1, padx=5, pady=5)
    cb_pac.set("Seleccionar paciente...")


    # tabs
    nb_frame = ctk.CTkFrame(v)
    nb_frame.pack(fill="both", expand=True, padx=10, pady=10)

    # usaremos ttk.Notebook
    from tkinter import ttk as ttk2
    notebook = ttk2.Notebook(nb_frame)
    notebook.pack(fill="both", expand=True)

    frame_turnos = ctk.CTkFrame(notebook)
    frame_trat = ctk.CTkFrame(notebook)
    frame_pagos = ctk.CTkFrame(notebook)
    frame_presup = ctk.CTkFrame(notebook)

    notebook.add(frame_turnos, text="Turnos")
    notebook.add(frame_trat, text="Tratamientos")
    notebook.add(frame_pagos, text="Pagos")
    notebook.add(frame_presup, text="Presupuestos")

    # tablas
    tabla_turnos = ttk.Treeview(
        frame_turnos,
        columns=("Fecha","Hora","Odontologo","Motivo","Estado"),
        show="headings"
    )
    for col in ("Fecha","Hora","Odontologo","Motivo","Estado"):
        tabla_turnos.heading(col, text=col)
        tabla_turnos.column(col, width=140)
    tabla_turnos.pack(fill="both", expand=True, padx=5, pady=5)

    tabla_trat = ttk.Treeview(
        frame_trat,
        columns=("Fecha","Odontologo","Tratamiento","Obs"),
        show="headings"
    )
    for col in ("Fecha","Odontologo","Tratamiento","Obs"):
        tabla_trat.heading(col, text=col)
        tabla_trat.column(col, width=160)
    tabla_trat.pack(fill="both", expand=True, padx=5, pady=5)

    tabla_pagos = ttk.Treeview(
        frame_pagos,
        columns=("Fecha","Monto","Metodo"),
        show="headings"
    )
    for col in ("Fecha","Monto","Metodo"):
        tabla_pagos.heading(col, text=col)
        tabla_pagos.column(col, width=160)
    tabla_pagos.pack(fill="both", expand=True, padx=5, pady=5)

    tabla_presup = ttk.Treeview(
        frame_presup,
        columns=("Fecha","Total","Obs"),
        show="headings"
    )
    for col in ("Fecha","Total","Obs"):
        tabla_presup.heading(col, text=col)
        tabla_presup.column(col, width=200)
    tabla_presup.pack(fill="both", expand=True, padx=5, pady=5)

    def cargar_historial():
        if cb_pac.get() == "":
            return
        pac_id = lista_pac[cb_pac.get()]

        # limpiar todas
        for t in (tabla_turnos, tabla_trat, tabla_pagos, tabla_presup):
            for f in t.get_children():
                t.delete(f)

        
        # tratamientos realizados
        filas = ejecutar_sql("""
            SELECT tr.fecha,
                   CONCAT(o.nombre,' ',o.apellido),
                   t.nombre,
                   tr.observaciones
            FROM tratamiento_realizado tr
            JOIN odontologo o ON o.odo_id = tr.odo_id
            JOIN tratamiento t ON t.tra_id = tr.tra_id
            WHERE tr.pac_id = %s
            ORDER BY tr.fecha
        """, (pac_id,))
        for fila in filas:
            tabla_trat.insert("", "end", values=fila)

        # pagos
        filas = ejecutar_sql("""
            SELECT fecha, monto, metodo
            FROM pago
            WHERE pac_id = %s
            ORDER BY fecha
        """, (pac_id,))
        for fila in filas:
            tabla_pagos.insert("", "end", values=fila)

        # presupuestos
        filas = ejecutar_sql("""
            SELECT fecha, total, observacion
            FROM presupuesto
            WHERE pac_id = %s
            ORDER BY fecha
        """, (pac_id,))
        for fila in filas:
            tabla_presup.insert("", "end", values=fila)

    ctk.CTkButton(frame_top, text="Cargar historial", width=150,
                  command=cargar_historial).grid(row=0, column=2, padx=5, pady=5)

# ---------- MODULO PRESUPUESTOS ----------
def ventana_presupuestos():
    v = ctk.CTkToplevel()
    v.title("Presupuestos")
    v.geometry("900x500")

    ctk.CTkLabel(v, text="Presupuesto de tratamientos",
                 font=("Arial", 20, "bold")).pack(pady=10)

    frame_top = ctk.CTkFrame(v)
    frame_top.pack(pady=5, padx=10, fill="x")

    # pacientes
    pacientes = ejecutar_sql("SELECT pac_id, CONCAT(nombre,' ',apellido) FROM paciente")
    lista_pac = {n: i for i, n in pacientes}

    ctk.CTkLabel(frame_top, text="Paciente:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
    cb_pac = ctk.CTkComboBox(frame_top, values=list(lista_pac.keys()), width=300)
    cb_pac.grid(row=0, column=1, padx=5, pady=5, sticky="w")

    # tratamientos
    tratamientos = ejecutar_sql("SELECT tra_id, nombre, precio FROM tratamiento")
    lista_tra = {}
    nombres_tra = []
    for tra_id, nombre, precio in tratamientos:
        lista_tra[nombre] = (tra_id, float(precio))
        nombres_tra.append(nombre)

    ctk.CTkLabel(frame_top, text="Tratamiento:").grid(row=1, column=0, padx=5, pady=5, sticky="e")
    cb_tra = ctk.CTkComboBox(frame_top, values=nombres_tra, width=300)
    cb_tra.grid(row=1, column=1, padx=5, pady=5, sticky="w")

    ctk.CTkLabel(frame_top, text="Cantidad:").grid(row=2, column=0, padx=5, pady=5, sticky="e")
    entry_cant = ctk.CTkEntry(frame_top, width=80, placeholder_text="1")
    entry_cant.grid(row=2, column=1, padx=5, pady=5, sticky="w")

    # tabla detalle
    frame_tabla = ctk.CTkFrame(v)
    frame_tabla.pack(pady=10, padx=10, fill="both", expand=True)

    tabla = ttk.Treeview(
        frame_tabla,
        columns=("Tratamiento","Cantidad","Precio","Subtotal"),
        show="headings"
    )
    for col in ("Tratamiento","Cantidad","Precio","Subtotal"):
        tabla.heading(col, text=col)
        tabla.column(col, width=150)
    tabla.pack(fill="both", expand=True, padx=5, pady=5)

    total_var = ctk.StringVar(value="0.00")
    ctk.CTkLabel(v, text="Total:").pack()
    lbl_total = ctk.CTkLabel(v, textvariable=total_var, font=("Arial", 18, "bold"))
    lbl_total.pack(pady=5)

    detalle = []  # lista de (tra_id, nombre, cant, precio, subtotal)

    def actualizar_total():
        t = sum(d[4] for d in detalle)
        total_var.set(f"{t:.2f}")

    def agregar_detalle():
        if cb_tra.get() == "":
            messagebox.showwarning("Atencion", "Selecciona un tratamiento")
            return
        try:
            cant = int(entry_cant.get() or "1")
        except:
            messagebox.showwarning("Atencion", "Cantidad invalida")
            return
        tra_id, precio = lista_tra[cb_tra.get()]
        subt = precio * cant
        detalle.append((tra_id, cb_tra.get(), cant, precio, subt))
        tabla.insert("", "end", values=(cb_tra.get(), cant, f"{precio:.2f}", f"{subt:.2f}"))
        actualizar_total()

    def guardar_presupuesto():
        if cb_pac.get() == "":
            messagebox.showwarning("Atencion", "Selecciona un paciente")
            return
        if not detalle:
            messagebox.showwarning("Atencion", "No hay tratamientos en el presupuesto")
            return

        pac_id = lista_pac[cb_pac.get()]
        total = sum(d[4] for d in detalle)

        # insertar presupuesto
        ejecutar_sql(
            "INSERT INTO presupuesto (pac_id, fecha, total, observacion) "
            "VALUES (%s, CURDATE(), %s, %s)",
            (pac_id, total, ""), commit=True
        )
        # obtener el ultimo id
        fila = ejecutar_sql("SELECT MAX(pre_id) FROM presupuesto")
        pre_id = fila[0][0]

        # insertar detalle
        for tra_id, nombre, cant, precio, subt in detalle:
            ejecutar_sql(
                "INSERT INTO presupuesto_detalle (pre_id, tra_id, cantidad, subtotal) "
                "VALUES (%s,%s,%s,%s)",
                (pre_id, tra_id, cant, subt), commit=True
            )

        messagebox.showinfo("Exito", "Presupuesto guardado correctamente")
        tabla.delete(*tabla.get_children())
        detalle.clear()
        actualizar_total()

    btns = ctk.CTkFrame(v)
    btns.pack(pady=5)
    ctk.CTkButton(btns, text="Agregar", width=120,
                  command=agregar_detalle).grid(row=0, column=0, padx=5)
    ctk.CTkButton(btns, text="Guardar presupuesto", width=180,
                  command=guardar_presupuesto).grid(row=0, column=1, padx=5)

# ---------- MODULO FICHA CLINICA ----------
def ventana_ficha_clinica():
    v = ctk.CTkToplevel()
    v.title("Ficha clínica del paciente")
    v.geometry("700x400")

    ctk.CTkLabel(v, text="Ficha clínica del paciente",
                 font=("Arial", 20, "bold")).pack(pady=10)

    frame = ctk.CTkFrame(v)
    frame.pack(pady=10, padx=10, fill="x")

    # lista de pacientes
    pacientes = ejecutar_sql("SELECT pac_id, CONCAT(nombre,' ',apellido) FROM paciente")
    lista_pac = {n: i for i, n in pacientes}

    ctk.CTkLabel(frame, text="Paciente:").grid(row=0, column=0, pady=5, padx=5, sticky="e")
    cb_pac = ctk.CTkComboBox(frame, values=list(lista_pac.keys()), width=300)
    cb_pac.grid(row=0, column=1, pady=5, padx=5)

    ctk.CTkLabel(frame, text="Enfermedades previas:").grid(row=1, column=0, sticky="ne", padx=5, pady=5)
    txt_enf = ctk.CTkTextbox(frame, width=400, height=60)
    txt_enf.grid(row=1, column=1, pady=5, padx=5)

    ctk.CTkLabel(frame, text="Alergias:").grid(row=2, column=0, sticky="ne", padx=5, pady=5)
    txt_ale = ctk.CTkTextbox(frame, width=400, height=60)
    txt_ale.grid(row=2, column=1, pady=5, padx=5)

    ctk.CTkLabel(frame, text="Medicacion actual:").grid(row=3, column=0, sticky="ne", padx=5, pady=5)
    txt_med = ctk.CTkTextbox(frame, width=400, height=60)
    txt_med.grid(row=3, column=1, pady=5, padx=5)

    ctk.CTkLabel(frame, text="Observaciones:").grid(row=4, column=0, sticky="ne", padx=5, pady=5)
    txt_obs = ctk.CTkTextbox(frame, width=400, height=60)
    txt_obs.grid(row=4, column=1, pady=5, padx=5)

    def limpiar():
        txt_enf.delete("1.0", "end")
        txt_ale.delete("1.0", "end")
        txt_med.delete("1.0", "end")
        txt_obs.delete("1.0", "end")

    def cargar_ficha(*args):
        limpiar()
        if cb_pac.get() == "":
            return
        pac_id = lista_pac[cb_pac.get()]
        filas = ejecutar_sql(
            "SELECT enfermedades_previas, alergias, medicacion_actual, observaciones "
            "FROM historia_medica WHERE pac_id=%s", (pac_id,)
        )
        if filas:
            enf, ale, med, obs = filas[0]
            if enf: txt_enf.insert("1.0", enf)
            if ale: txt_ale.insert("1.0", ale)
            if med: txt_med.insert("1.0", med)
            if obs: txt_obs.insert("1.0", obs)

    def guardar_ficha():
        if cb_pac.get() == "":
            messagebox.showwarning("Atencion", "Debes seleccionar un paciente")
            return

        pac_id = lista_pac[cb_pac.get()]
        enf = txt_enf.get("1.0", "end").strip()
        ale = txt_ale.get("1.0", "end").strip()
        med = txt_med.get("1.0", "end").strip()
        obs = txt_obs.get("1.0", "end").strip()

        filas = ejecutar_sql("SELECT his_id FROM historia_medica WHERE pac_id=%s", (pac_id,))
        if filas:
            # actualizar
            ejecutar_sql(
                "UPDATE historia_medica SET enfermedades_previas=%s, alergias=%s, "
                "medicacion_actual=%s, observaciones=%s WHERE pac_id=%s",
                (enf, ale, med, obs, pac_id), commit=True
            )
            messagebox.showinfo("Exito", "Ficha clínica actualizada")
        else:
            # insertar
            ejecutar_sql(
                "INSERT INTO historia_medica (pac_id, enfermedades_previas, alergias, "
                "medicacion_actual, observaciones) VALUES (%s,%s,%s,%s,%s)",
                (pac_id, enf, ale, med, obs), commit=True
            )
            messagebox.showinfo("Exito", "Ficha clínica guardada")

    cb_pac.configure(command=cargar_ficha)

    ctk.CTkButton(v, text="Guardar ficha", width=200,
                  command=guardar_ficha).pack(pady=10)




# ---------- MODULO PAGOS ----------
def ventana_pagos():
    v = ctk.CTkToplevel()
    v.title("Pagos")
    v.geometry("600x400")
    ctk.CTkLabel(v, text="Registro de Pagos", font=("Arial", 20, "bold")).pack(pady=10)

    frame = ctk.CTkFrame(v)
    frame.pack(pady=10)

    pacientes = ejecutar_sql("SELECT pac_id, CONCAT(nombre,' ',apellido) FROM paciente")
    lista_pac = {f"{n}": i for i,n in pacientes}

    cb_pac = ctk.CTkComboBox(frame, values=list(lista_pac.keys()), width=250)
    monto = ctk.CTkEntry(frame, width=150, placeholder_text="Monto")
    metodo = ctk.CTkComboBox(frame, values=["Efectivo","Tarjeta","Transferencia"], width=150)

    ctk.CTkLabel(frame, text="Paciente:").grid(row=0, column=0)
    cb_pac.grid(row=0, column=1)
    ctk.CTkLabel(frame, text="Monto:").grid(row=1, column=0)
    monto.grid(row=1, column=1)
    ctk.CTkLabel(frame, text="Método:").grid(row=2, column=0)
    metodo.grid(row=2, column=1)

    def guardar_pago():
        sql = "INSERT INTO pago (pac_id, fecha, monto, metodo) VALUES (%s,%s,%s,%s)"
        datos = (lista_pac[cb_pac.get()], date.today(), monto.get(), metodo.get())
        ejecutar_sql(sql, datos, commit=True)
        messagebox.showinfo("Éxito", "Pago registrado correctamente")
        cargar_tabla()

    ctk.CTkButton(frame, text="Guardar", command=guardar_pago).grid(row=3, column=0, columnspan=2, pady=10)

    tabla = ttk.Treeview(v, columns=("ID","Paciente","Monto","Método","Fecha"), show="headings")
    for col in ("ID","Paciente","Monto","Método","Fecha"):
        tabla.heading(col, text=col)
    tabla.pack(fill="x", pady=10)

    def cargar_tabla():
        for f in tabla.get_children():
            tabla.delete(f)
        filas = ejecutar_sql("""SELECT p.pag_id, pa.nombre, p.monto, p.metodo, p.fecha
                             FROM pago p JOIN paciente pa ON p.pac_id=pa.pac_id""")
        for fila in filas:
            tabla.insert("", "end", values=fila)

    cargar_tabla()

def reporte_pagos():
    filas = ejecutar_sql("""
        SELECT p.pag_id,
               CONCAT(pa.nombre,' ',pa.apellido) AS paciente,
               p.fecha, p.monto, p.metodo
        FROM pago p
        JOIN paciente pa ON pa.pac_id = p.pac_id
        ORDER BY p.fecha DESC
    """)
    headers = ["ID", "Paciente", "Fecha", "Monto", "Metodo"]
    exportar_excel("reporte_pagos", headers, filas)


# ---------- VENTANA PRINCIPAL ----------
app = ctk.CTk()
app.title("Clínica Dental - Sistema")
app.geometry("700x800")
app.minsize(650, 750)
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("green")
# logo
logo_img = ctk.CTkImage(
    light_image=Image.open("logo.png.jpeg"),
    size=(280, 120)
)

logo = ctk.CTkLabel(app, image=logo_img, text="")
logo.pack(pady=(20, 10))

titulo = ctk.CTkLabel(
    app,
    text="Sistema de Gestión Odontológica",
    font=("Segoe UI", 20, "bold"),
    text_color="#1C7CB8"
)
titulo.pack(pady=(0, 25))

def boton_principal(texto, comando):
    return ctk.CTkButton(
        app,
        text=texto,
        width=320,
        height=45,
        corner_radius=12,
        font=("Segoe UI", 15, "bold"),
        fg_color="#2EA8E5",
        hover_color="#1C7CB8",
        command=comando
    )
boton_principal("Pacientes", ventana_pacientes).pack(pady=8)
boton_principal("Odontólogos", ventana_odontologos).pack(pady=8)
boton_principal("Turnos", ventana_turnos).pack(pady=8)
boton_principal("Tratamientos", ventana_tratamientos).pack(pady=8)
boton_principal("Pagos", ventana_pagos).pack(pady=8)
boton_principal("Ficha clínica", ventana_ficha_clinica).pack(pady=8)
boton_principal("Presupuestos", ventana_presupuestos).pack(pady=8)
boton_principal(
    "Reporte General (Excel)",
    reporte_general_excel
).pack(pady=8)


#color del boton salir
ctk.CTkButton(
    app,
    text="Salir",
    width=320,
    height=45,
    corner_radius=12,
    font=("Segoe UI", 15, "bold"),
    fg_color="#7AC943",
    hover_color="#5AAE2E",
    command=app.destroy
).pack(pady=(20, 10))
#fondo
ctk.set_appearance_mode("light")
app.configure(fg_color="#F4F9FC")


app.mainloop()
